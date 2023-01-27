import PyPDF2
import openpyxl

reader = PyPDF2.PdfReader('IITM_Faculty_Expertise.pdf')
excel = openpyxl.load_workbook('FacultyUploadTemplate.xlsx')
sheet = excel['Faculty Member']

r = 2
c = 1

# code for custom pages
# start = int(input("Enter starting page no: "))
# end = int(input("Enter ending page no: "))

start = 1
end = len(reader.pages)

for i in range(start - 1, end):
  page = reader.pages[i]
  entire = page.extract_text()
  lines = entire.splitlines()

  try:
    if not lines[1].startswith("PhD"):
      continue
  except:
    continue

  no = 0

  # for name and title
  
  line = lines[no]
  s = line.split('.')
  if len(s) != 1:
    title = sheet.cell(r, c)
    title.value = s[0]
    c += 1

    name = sheet.cell(r, c)
    name.value = s[1]
    c += 1

  else:
    c += 1
    name = sheet.cell(r, c)
    name.value = line
    c += 1

  no += 1
  if lines[no].startswith('PhD'):
    no += 1

  # for Designation and Department
  
  try:
    line = lines[no]
    s = line.split(',')
    designation = sheet.cell(r, c)
    designation.value = s[0]
    c += 1
    
    department = sheet.cell(r, c)
    department.value = s[1]
    c += 1
  except:
    c += 1

  no += 1

  # for email and office phone number
  try:
    line = lines[no]
    s = line.split(';')
    email = sheet.cell(r, c)
    email.value = s[1]
    c += 1
  except:
    c += 1


  phone = sheet.cell(r, c)
  phone.value = s[0]
  c += 1


  no += 1

  # for mobile
  c += 1

  try:
    if lines[no].startswith("http"):#skip homepage for now.
      no += 1
  except:
    pass
  try:
    if lines[no].startswith("http"):#in some cases 2 homepages present.
      no += 1
  except:
    pass

  # for specialization
  try:
    if lines[no].startswith("Back"):#avoid recurring "back to top"
      no += 1
    specialization = sheet.cell(r, c)
    specialization.value = '\n'.join(lines[no:no + 2])
    c += 1
  except:
    c += 1

  no -= 1

  # for homepage
  try:
    if lines[no].startswith("http"):#resuming for homepage/website
      homepage = sheet.cell(r, c)
      homepage.value = lines[4]
    c += 1
  except:
    c += 1

  # for website
    c += 1 
  

  r += 1 #next excel row
  c = 1  #start column from 1 


excel.save('FacultyUploadTemplate1.xlsx')

